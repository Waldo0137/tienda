import json
import sys
import io
import uuid
from collections import deque
from datetime import datetime, date
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.db import transaction
from django.db.models import Count, Sum
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment
from xhtml2pdf import pisa

from inventory.models import *
from pos.models import *
from django.views.generic import ListView, FormView
from report.forms import ReportForm, YearReportForm, MonthReportForm, DayReportForm


MONTH_NAMES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def is_leap_year(year):
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)

def is_valid_day(year, month, day):
    if month in [4, 6, 9, 11]:
        return day <= 30
    elif month == 2:
        if is_leap_year(year):
            return day <= 29
        else:
            return day <= 28
    else:
        return day <= 31


class SalesListView(ListView):
    model = Sales
    template_name = 'report/sales_list.html'
    context_object_name = 'sales'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_clientes'] = Sales.objects.values('id').distinct().count()
        context['total_items_vendidos'] = SalesItems.objects.aggregate(total=Sum('qty'))['total']
        context['total_ingresos'] = Sales.objects.aggregate(total=Sum('grand_total'))['total']
        return context
    
    
class GenerateExcelSalesView(View):
    def post(self, request, *args, **kwargs):
        form = ReportForm(request.POST)
        if form.is_valid():
            
            sales = Sales.objects.all()

            
            total_clientes = Sales.objects.values('id').distinct().count()

            
            total_items_vendidos = salesItems.objects.aggregate(total=Sum('qty'))['total']

            
            total_ingresos = Sales.objects.aggregate(total=Sum('grand_total'))['total']

            
            wb = Workbook()
            ws = wb.active

            
            headers = ['Cliente', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
            ws.append([''] + headers)

            
            for sale in sales:
                items = salesItems.objects.filter(sale=sale).all()
                products_list = {}
                for item in items:
                    product_name = item.product.name
                    if product_name in products_list:
                        products_list[product_name] += item.qty
                    else:
                        products_list[product_name] = item.qty

                total_items_sold = sum(products_list.values())

                # Formatear los datos según sea necesario
                sale_data = [
                    '',
                    sale.cliente,
                    sale.date_added.strftime('%Y-%m-%d %H:%M'),  
                    ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),  
                    sale.grand_total,
                    total_items_sold
                ]

                ws.append(sale_data)

            
            total_row = ['Total General:', total_clientes, '', '', total_ingresos, total_items_vendidos]
            ws.append(total_row)

            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.column in [2, 3, 4]:  
                        cell.alignment = Alignment(horizontal='left')
                    elif cell.column in [5, 6, 7]:  
                        cell.alignment = Alignment(horizontal='right')

            
            total_general_cell = ws['A' + str(ws.max_row)]
            total_general_cell.alignment = Alignment(horizontal='left')

            
            total_clientes_cell = ws['B' + str(ws.max_row)]
            total_clientes_cell.alignment = Alignment(horizontal='right')

            total_ingresos_cell = ws['E' + str(ws.max_row)]
            total_ingresos_cell.alignment = Alignment(horizontal='right')

            total_items_vendidos_cell = ws['F' + str(ws.max_row)]
            total_items_vendidos_cell.alignment = Alignment(horizontal='right')

            current_date = datetime.now()
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=reporte_ventas_general_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

            wb.save(response)

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")

class GenerateExcelSalesYearView(View):
    def post(self, request, *args, **kwargs):
        form = YearReportForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data.get('year')

            
            sales = Sales.objects.filter(date_added__year=year)

            
            total_clientes = Sales.objects.filter(date_added__year=year).values('id').distinct().count()

            
            total_items_vendidos = salesItems.objects.filter(sale__in=sales).aggregate(total=Sum('qty'))['total']
            total_ingresos = sales.aggregate(total=Sum('grand_total'))['total']

            
            wb = Workbook()
            ws = wb.active

            ws.append([f"Reporte de Ventas: del {year}"])
            
            headers = ['Cliente', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
            ws.append([''] + headers)

            
            for sale in sales:
                items = salesItems.objects.filter(sale=sale).all()
                products_list = {}
                for item in items:
                    product_name = item.product.name
                    if product_name in products_list:
                        products_list[product_name] += item.qty
                    else:
                        products_list[product_name] = item.qty

                total_items_sold = sum(products_list.values())

                
                sale_data = [
                    '',
                    sale.cliente,
                    sale.date_added.strftime('%Y-%m-%d %H:%M'),
                    ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                    sale.grand_total,
                    total_items_sold
                ]

                ws.append(sale_data)

            
            total_row = ['Total General:', total_clientes, '', '', total_ingresos, total_items_vendidos]
            ws.append(total_row)

            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.column in [2, 3, 4]: 
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column in [5, 6, 7]: 
                        cell.alignment = Alignment(horizontal='center')

            
            total_general_cell = ws['A' + str(ws.max_row)]
            total_general_cell.alignment = Alignment(horizontal='center')

            
            total_clientes_cell = ws['B' + str(ws.max_row)]
            total_clientes_cell.alignment = Alignment(horizontal='center')

            total_ingresos_cell = ws['E' + str(ws.max_row)]
            total_ingresos_cell.alignment = Alignment(horizontal='center')

            total_items_vendidos_cell = ws['F' + str(ws.max_row)]
            total_items_vendidos_cell.alignment = Alignment(horizontal='center')

            current_date = datetime.now()
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=reporte_ventas_anual_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

            wb.save(response)

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")

class GenerateExcelSalesMonthView(View):
    def post(self, request, *args, **kwargs):
        form = MonthReportForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data.get('year')
            month = form.cleaned_data.get('month')

            try:
                month_name = MONTH_NAMES[month - 1]
            except IndexError:
                return HttpResponseBadRequest("El mes proporcionado no es válido.")

            sales = Sales.objects.filter(date_added__year=year, date_added__month=month)
            total_clientes = sales.values('id').distinct().count()
            total_items_vendidos = salesItems.objects.filter(sale__in=sales).aggregate(total=Sum('qty'))['total']
            total_ingresos = sales.aggregate(total=Sum('grand_total'))['total']

            sale_details = []
            for sale in sales:
                items = salesItems.objects.filter(sale=sale)
                products_list = {}
                for item in items:
                    product_name = item.product.name
                    products_list[product_name] = products_list.get(product_name, 0) + item.qty
                sale_details.append({
                    'cliente': sale.cliente,
                    'date_added': sale.date_added,
                    'products_list': products_list,
                    'grand_total': sale.grand_total,
                    'total_items_sold': sum(products_list.values())
                })

            
            wb = Workbook()
            ws = wb.active
            ws.append([f"Reporte de Ventas: de {month_name} del {year}"])
            
            headers = ['Cliente', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
            ws.append([''] + headers)

            
            for sale in sale_details:
                products_list = sale['products_list']
                total_items_sold = sale['total_items_sold']

                sale_data = [
                    '',
                    sale['cliente'],
                    sale['date_added'].strftime('%Y-%m-%d %H:%M'),
                    ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                    sale['grand_total'],
                    total_items_sold
                ]
                ws.append(sale_data)

            
            total_row = ['Total General:', total_clientes, '', '', total_ingresos, total_items_vendidos]
            ws.append(total_row)

            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.column in [2, 3, 4]:
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column in [5, 6, 7]:
                        cell.alignment = Alignment(horizontal='center')

            # Alinear la celda "Total General:" a la izquierda
            total_general_cell = ws['A' + str(ws.max_row)]
            total_general_cell.alignment = Alignment(horizontal='center')

            # Alinear los totales a la derecha
            total_clientes_cell = ws['B' + str(ws.max_row)]
            total_clientes_cell.alignment = Alignment(horizontal='center')

            total_ingresos_cell = ws['E' + str(ws.max_row)]
            total_ingresos_cell.alignment = Alignment(horizontal='center')

            total_items_vendidos_cell = ws['F' + str(ws.max_row)]
            total_items_vendidos_cell.alignment = Alignment(horizontal='center')

            current_date = datetime.now()
            filename = f"reporte_ventas_mensual_{current_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")





class GenerateExcelSalesDayView(View):
    def post(self, request, *args, **kwargs):
        form = DayReportForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data.get('year')
            month = form.cleaned_data.get('month')
            day = form.cleaned_data.get('day')

            try:
                month_name = MONTH_NAMES[month - 1]
            except IndexError:
                messages.error(self.request, "El mes proporcionado no es válido.")
                return HttpResponseBadRequest("El mes proporcionado no es válido.")

            if not is_valid_day(year, month, day):
                messages.error(self.request, "La fecha ingresada no es válida.")
                return HttpResponseBadRequest("La fecha ingresada no es válida.")

            sales = Sales.objects.filter(date_added__year=year, date_added__month=month, date_added__day=day)
            total_clientes = sales.values('id').distinct().count()
            total_items_vendidos = salesItems.objects.filter(sale__in=sales).aggregate(total=Sum('qty'))['total']
            total_ingresos = sales.aggregate(total=Sum('grand_total'))['total']

            sale_details = []
            for sale in sales:
                items = salesItems.objects.filter(sale=sale)
                products_list = {}
                for item in items:
                    product_name = item.product.name
                    products_list[product_name] = products_list.get(product_name, 0) + item.qty
                sale_details.append({
                    'cliente': sale.cliente,
                    'date_added': sale.date_added,
                    'products_list': products_list,
                    'grand_total': sale.grand_total,
                    'total_items_sold': sum(products_list.values())
                })

            
            wb = Workbook()
            ws = wb.active

            ws.append([f"Reporte de Ventas - Día: {day} de {month_name} del {year}"])
            
            
            headers = ['Cliente', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
            ws.append(['']+ headers)

            
            for sale in sale_details:
                products_list = sale['products_list']
                sale_data = [
                    '',
                    sale['cliente'],
                    sale['date_added'].strftime('%Y-%m-%d %H:%M'),
                    ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                    sale['grand_total'],
                    sale['total_items_sold']
                ]
                ws.append(sale_data)

            
            total_row = ['Total General:', total_clientes, '','', total_ingresos, total_items_vendidos]
            ws.append(total_row)

            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

            current_date = datetime.now()
            filename = f"reporte_ventas_diario_{current_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")
