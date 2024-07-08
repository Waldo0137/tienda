from pos.models import Sales
from purchase.models import PurchaseProduct
from report.forms import *
from datetime import datetime
from decimal import Decimal
from io import BytesIO
import uuid

from django.http import HttpResponse, HttpResponseBadRequest
from django.views.generic import FormView
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.db.models import Sum

from django.views import View  
class GenerateExcelProfitView(View):
    def post(self, request, *args, **kwargs):
        form = SalesReportForm(request.POST)
        if form.is_valid():
            sales_queryset = self.get_queryset(form)

            total_ingresos = self.calculate_total_ingresos(sales_queryset)
            total_costos = self.calculate_total_costos(sales_queryset)
            total_ingresos_decimal = Decimal(total_ingresos)
            total_costos_decimal = Decimal(total_costos)
            total_ganancia = total_ingresos_decimal - total_costos_decimal

            sales_data, total_utilidades = self.get_sales_data_and_utilities(sales_queryset)

            current_date = datetime.now()
            username = request.user.username
            unique_key = str(uuid.uuid4())

            
            excel_file = self.generate_excel_file(sales_data)

            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="reporte_ganancias_general_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx"'
            response.write(excel_file.getvalue())

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")

    def get_queryset(self, form):
        queryset = Sales.objects.all()

        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            if start_date:
                queryset = queryset.filter(date_added__gte=start_date)
            if end_date:
                queryset = queryset.filter(date_added__lte=end_date)

        return queryset

    def calculate_total_ingresos(self, sales_queryset):
        total_ingresos = sales_queryset.aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
        return total_ingresos

    def calculate_total_costos(self, sales_queryset):
        total_costos = Decimal('0')

        for sale in sales_queryset:
            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    costo_producto = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    costo_total = costo_producto * Decimal(qty_comprada)
                    total_costos += costo_total

        return total_costos

    def get_sales_data_and_utilities(self, sales_queryset):
        sales_data = []
        total_utilidades = Decimal(0)

        for sale in sales_queryset:
            sale_cost = self.calculate_sale_cost(sale)
            sale_profit = Decimal(sale.grand_total) - sale_cost
            products_list = []

            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    cost_per_unit = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    total_qty_vendida = item.qty
                    total_qty_comprada = qty_comprada

                    product_ganancia = (Decimal(item.qty) * sale_profit) / Decimal(total_qty_vendida)

                    total_gasto_compras = cost_per_unit * Decimal(total_qty_comprada)

                    ganancia_bruta = (sale_cost + product_ganancia) - total_gasto_compras
                    total_utilidades += ganancia_bruta
                    products_list.append({
                        'product_name': item.product.name,
                        'cost_per_unit': cost_per_unit,
                        'total_qty_vendida': total_qty_vendida,
                        'total_qty_comprada': total_qty_comprada,
                        'product_ganancia': product_ganancia,
                        'ganancia_estado': 'Positiva' if product_ganancia > 0 else ('Negativa' if product_ganancia < 0 else 'Neutra'),
                        'total_gasto_compras': total_gasto_compras,
                        'ganancia_bruta': ganancia_bruta,
                    })

            sales_data.append({
                'date_added': sale.date_added,
                'products_list': products_list,
                'venta_total': Decimal(sale.grand_total),
                'costo_total': sale_cost,
                'ganancia_total': sale_profit,
            })

        return sales_data, total_utilidades

    def calculate_sale_cost(self, sale):
        sale_cost = Decimal('0')
        for item in sale.salesitems_set.all():
            purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
            if purchase_product:
                costo_producto = purchase_product.cost
                sale_cost += costo_producto * Decimal(item.qty)
        return sale_cost

    def generate_excel_file(self, sales_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ganancias"

        # Encabezado
        headers = [
            "Fecha de Venta", "Nombre del Producto", "Costo por Unidad", "Cantidad Vendida",
            "Cantidad Comprada", "Ganancia por Producto", "Estado de Ganancia", "Total de Gasto en Compras",
            "Ganancia Bruta"
        ]
        ws.append(headers)

        # Datos
        for sale in sales_data:
            for product in sale['products_list']:
                row = [
                    sale['date_added'].strftime('%Y-%m-%d %H:%M:%S'),
                    product['product_name'],
                    product['cost_per_unit'],
                    product['total_qty_vendida'],
                    product['total_qty_comprada'],
                    product['product_ganancia'],
                    product['ganancia_estado'],
                    product['total_gasto_compras'],
                    product['ganancia_bruta']
                ]
                ws.append(row)

        # Ajustar el ancho de las columnas
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width


        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file


class YearlyExcelProfitView(View):
    def post(self, request, *args, **kwargs):
        form = YearReportForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data.get('year')
            sales_queryset = self.get_queryset(year)

            total_ingresos = self.calculate_total_ingresos(sales_queryset)
            total_costos = self.calculate_total_costos(sales_queryset)
            total_ingresos_decimal = Decimal(total_ingresos)
            total_costos_decimal = Decimal(total_costos)
            total_ganancia = total_ingresos_decimal - total_costos_decimal

            sales_data, total_utilidades = self.get_sales_data_and_utilities(sales_queryset)

            current_date = datetime.now()
            username = request.user.username
            unique_key = str(uuid.uuid4())


            excel_file = self.generate_excel_file(sales_data)


            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="reporte_ganancias_anual_{year}_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx"'
            response.write(excel_file.getvalue())

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")

    def get_queryset(self, year):
        return Sales.objects.filter(date_added__year=year)

    def calculate_total_ingresos(self, sales_queryset):
        total_ingresos = sales_queryset.aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
        return total_ingresos

    def calculate_total_costos(self, sales_queryset):
        total_costos = Decimal('0')

        for sale in sales_queryset:
            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    costo_producto = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    costo_total = costo_producto * Decimal(qty_comprada)
                    total_costos += costo_total

        return total_costos

    def get_sales_data_and_utilities(self, sales_queryset):
        sales_data = []
        total_utilidades = Decimal(0)

        for sale in sales_queryset:
            sale_cost = self.calculate_sale_cost(sale)
            sale_profit = Decimal(sale.grand_total) - sale_cost
            products_list = []

            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    cost_per_unit = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    total_qty_vendida = item.qty
                    total_qty_comprada = qty_comprada

                    product_ganancia = (Decimal(item.qty) * sale_profit) / Decimal(total_qty_vendida)

                    total_gasto_compras = cost_per_unit * Decimal(total_qty_comprada)

                    ganancia_bruta = (sale_cost + product_ganancia) - total_gasto_compras
                    total_utilidades += ganancia_bruta
                    products_list.append({
                        'product_name': item.product.name,
                        'cost_per_unit': cost_per_unit,
                        'total_qty_vendida': total_qty_vendida,
                        'total_qty_comprada': total_qty_comprada,
                        'product_ganancia': product_ganancia,
                        'ganancia_estado': 'Positiva' if product_ganancia > 0 else ('Negativa' if product_ganancia < 0 else 'Neutra'),
                        'total_gasto_compras': total_gasto_compras,
                        'ganancia_bruta': ganancia_bruta,
                    })

            sales_data.append({
                'date_added': sale.date_added,
                'products_list': products_list,
                'venta_total': Decimal(sale.grand_total),
                'costo_total': sale_cost,
                'ganancia_total': sale_profit,
            })

        return sales_data, total_utilidades

    def calculate_sale_cost(self, sale):
        sale_cost = Decimal('0')
        for item in sale.salesitems_set.all():
            purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
            if purchase_product:
                costo_producto = purchase_product.cost
                sale_cost += costo_producto * Decimal(item.qty)
        return sale_cost

    def generate_excel_file(self, sales_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ganancias"

    
        headers = [
            "Fecha de Venta", "Nombre del Producto", "Costo por Unidad", "Cantidad Vendida",
            "Cantidad Comprada", "Ganancia por Producto", "Estado de Ganancia", "Total de Gasto en Compras",
            "Ganancia Bruta"
        ]
        ws.append(headers)

    
        for sale in sales_data:
            for product in sale['products_list']:
                row = [
                    sale['date_added'].strftime('%Y-%m-%d %H:%M:%S'),
                    product['product_name'],
                    product['cost_per_unit'],
                    product['total_qty_vendida'],
                    product['total_qty_comprada'],
                    product['product_ganancia'],
                    product['ganancia_estado'],
                    product['total_gasto_compras'],
                    product['ganancia_bruta']
                ]
                ws.append(row)

    
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file


class MonthlyExcelProfitView(FormView):
    def post(self, request, *args, **kwargs):
        form = MonthYearReportForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data.get('year')
            month = form.cleaned_data.get('month')
            sales_queryset = self.get_queryset(year, month)

            total_ingresos = self.calculate_total_ingresos(sales_queryset)
            total_costos = self.calculate_total_costos(sales_queryset)
            total_ingresos_decimal = Decimal(total_ingresos)
            total_costos_decimal = Decimal(total_costos)
            total_ganancia = total_ingresos_decimal - total_costos_decimal

            sales_data, total_utilidades = self.get_sales_data_and_utilities(sales_queryset)

            current_date = datetime.now()
            username = request.user.username
            unique_key = str(uuid.uuid4())

        
            excel_file = self.generate_excel_file(sales_data)

        
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="reporte_ganancias_mensual_{month}_{year}_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx"'
            response.write(excel_file.getvalue())

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")

    def get_queryset(self, year, month):
        return Sales.objects.filter(date_added__year=year, date_added__month=month)

    def calculate_total_ingresos(self, sales_queryset):
        total_ingresos = sales_queryset.aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
        return total_ingresos

    def calculate_total_costos(self, sales_queryset):
        total_costos = Decimal('0')

        for sale in sales_queryset:
            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    costo_producto = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    costo_total = costo_producto * Decimal(qty_comprada)
                    total_costos += costo_total

        return total_costos

    def get_sales_data_and_utilities(self, sales_queryset):
        sales_data = []
        total_utilidades = Decimal(0)

        for sale in sales_queryset:
            sale_cost = self.calculate_sale_cost(sale)
            sale_profit = Decimal(sale.grand_total) - sale_cost
            products_list = []

            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    cost_per_unit = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    total_qty_vendida = item.qty
                    total_qty_comprada = qty_comprada

                    product_ganancia = (Decimal(item.qty) * sale_profit) / Decimal(total_qty_vendida)

                    total_gasto_compras = cost_per_unit * Decimal(total_qty_comprada)

                    ganancia_bruta = (sale_cost + product_ganancia) - total_gasto_compras
                    total_utilidades += ganancia_bruta
                    products_list.append({
                        'product_name': item.product.name,
                        'cost_per_unit': cost_per_unit,
                        'total_qty_vendida': total_qty_vendida,
                        'total_qty_comprada': total_qty_comprada,
                        'product_ganancia': product_ganancia,
                        'ganancia_estado': 'Positiva' if product_ganancia > 0 else ('Negativa' if product_ganancia < 0 else 'Neutra'),
                        'total_gasto_compras': total_gasto_compras,
                        'ganancia_bruta': ganancia_bruta,
                    })

            sales_data.append({
                'date_added': sale.date_added,
                'products_list': products_list,
                'venta_total': Decimal(sale.grand_total),
                'costo_total': sale_cost,
                'ganancia_total': sale_profit,
            })

        return sales_data, total_utilidades

    def calculate_sale_cost(self, sale):
        sale_cost = Decimal('0')
        for item in sale.salesitems_set.all():
            purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
            if purchase_product:
                costo_producto = purchase_product.cost
                sale_cost += costo_producto * Decimal(item.qty)
        return sale_cost

    def generate_excel_file(self, sales_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ganancias"

        
        headers = [
            "Fecha de Venta", "Nombre del Producto", "Costo por Unidad", "Cantidad Vendida",
            "Cantidad Comprada", "Ganancia por Producto", "Estado de Ganancia", "Total de Gasto en Compras",
            "Ganancia Bruta"
        ]
        ws.append(headers)

        
        for sale in sales_data:
            for product in sale['products_list']:
                row = [
                    sale['date_added'].strftime('%Y-%m-%d %H:%M:%S'),
                    product['product_name'],
                    product['cost_per_unit'],
                    product['total_qty_vendida'],
                    product['total_qty_comprada'],
                    product['product_ganancia'],
                    product['ganancia_estado'],
                    product['total_gasto_compras'],
                    product['ganancia_bruta']
                ]
                ws.append(row)


        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width


        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file


class DailyExcelProfitView(FormView):
    template_name = 'your_template.html'
    form_class = DayMonthYearReportForm

    def form_valid(self, form):
        year = form.cleaned_data.get('year')
        month = form.cleaned_data.get('month')
        day = form.cleaned_data.get('day')

        sales_queryset = self.get_queryset(year, month, day)

        total_ingresos = self.calculate_total_ingresos(sales_queryset)
        total_costos = self.calculate_total_costos(sales_queryset)
        total_ingresos_decimal = Decimal(total_ingresos)
        total_costos_decimal = Decimal(total_costos)
        total_ganancia = total_ingresos_decimal - total_costos_decimal

        sales_data, total_utilidades = self.get_sales_data_and_utilities(sales_queryset)

        current_date = datetime.now()
        username = self.request.user.username
        unique_key = str(uuid.uuid4())

        
        excel_file = self.generate_excel_file(sales_data)

        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_ganancias_diaria_{day}_{month}_{year}_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx"'
        response.write(excel_file.getvalue())

        return response

    def get_queryset(self, year, month, day):
        queryset = Sales.objects.filter(date_added__year=year, date_added__month=month)

        if day:
            queryset = queryset.filter(date_added__day=day)

        return queryset

    def calculate_total_ingresos(self, sales_queryset):
        total_ingresos = sales_queryset.aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
        return total_ingresos

    def calculate_total_costos(self, sales_queryset):
        total_costos = Decimal('0')

        for sale in sales_queryset:
            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    costo_producto = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    costo_total = costo_producto * Decimal(qty_comprada)
                    total_costos += costo_total

        return total_costos

    def get_sales_data_and_utilities(self, sales_queryset):
        sales_data = []
        total_utilidades = Decimal(0)

        for sale in sales_queryset:
            sale_cost = self.calculate_sale_cost(sale)
            sale_profit = Decimal(sale.grand_total) - sale_cost
            products_list = []

            for item in sale.salesitems_set.all():
                purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
                if purchase_product:
                    cost_per_unit = purchase_product.cost
                    qty_comprada = sum([pp.qty for pp in PurchaseProduct.objects.filter(product=item.product)])
                    total_qty_vendida = item.qty
                    total_qty_comprada = qty_comprada

                    product_ganancia = (Decimal(item.qty) * sale_profit) / Decimal(total_qty_vendida)

                    total_gasto_compras = cost_per_unit * Decimal(total_qty_comprada)

                    ganancia_bruta = (sale_cost + product_ganancia) - total_gasto_compras
                    total_utilidades += ganancia_bruta
                    products_list.append({
                        'product_name': item.product.name,
                        'cost_per_unit': cost_per_unit,
                        'total_qty_vendida': total_qty_vendida,
                        'total_qty_comprada': total_qty_comprada,
                        'product_ganancia': product_ganancia,
                        'ganancia_estado': 'Positiva' if product_ganancia > 0 else ('Negativa' if product_ganancia < 0 else 'Neutra'),
                        'total_gasto_compras': total_gasto_compras,
                        'ganancia_bruta': ganancia_bruta,
                    })

            sales_data.append({
                'date_added': sale.date_added,
                'products_list': products_list,
                'venta_total': Decimal(sale.grand_total),
                'costo_total': sale_cost,
                'ganancia_total': sale_profit,
            })

        return sales_data, total_utilidades

    def calculate_sale_cost(self, sale):
        sale_cost = Decimal('0')
        for item in sale.salesitems_set.all():
            purchase_product = PurchaseProduct.objects.filter(product=item.product).first()
            if purchase_product:
                costo_producto = purchase_product.cost
                sale_cost += costo_producto * Decimal(item.qty)
        return sale_cost

    def generate_excel_file(self, sales_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ganancias"

        
        headers = [
            "Fecha de Venta", "Nombre del Producto", "Costo por Unidad", "Cantidad Vendida",
            "Cantidad Comprada", "Ganancia por Producto", "Estado de Ganancia", "Total de Gasto en Compras",
            "Ganancia Bruta"
        ]
        ws.append(headers)

        
        for sale in sales_data:
            for product in sale['products_list']:
                row = [
                    sale['date_added'].strftime('%Y-%m-%d %H:%M:%S'),
                    product['product_name'],
                    product['cost_per_unit'],
                    product['total_qty_vendida'],
                    product['total_qty_comprada'],
                    product['product_ganancia'],
                    product['ganancia_estado'],
                    product['total_gasto_compras'],
                    product['ganancia_bruta']
                ]
                ws.append(row)

        
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file
