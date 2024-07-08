from django.views import View
from django.http import HttpResponse, HttpResponseBadRequest
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from django.db.models import Sum
from purchase.models import PurchaseProduct
from report.forms import ReportForm
from django.views.generic.edit import FormView
from django.http import HttpResponse, HttpResponseBadRequest
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from purchase.models import PurchaseProduct
from report.forms import YearReportForm
from report.forms import DayMonthYearReportForm
from report.forms import MonthYearReportForm

MONTH_NAMES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def is_leap_year(year):
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)

def is_valid_day(year, month, day):
    if month in [4, 6, 9, 11]:
        return day <= 30
    elif month == 2:
        if is_leap_year(year):
            return day <= 29
        else:
            return day <= 28
    else:
        return day <= 31

class GenerateExcelPurchaseView(View):
    def post(self, request, *args, **kwargs):
        form = ReportForm(request.POST)
        if form.is_valid():
            purchase_products = PurchaseProduct.objects.all()

            
            total_suppliers = purchase_products.values('supplier').distinct().count()

            
            total_items_comprados = purchase_products.aggregate(total=Sum('qty'))['total']

            
            total_costos = purchase_products.aggregate(total=Sum('total'))['total']

            
            wb = Workbook()
            ws = wb.active

            
            headers = ['Proveedor', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
            ws.append([''] + headers)

            
            for purchase_product in purchase_products:
                products_list = {purchase_product.product.name: purchase_product.qty}
                total_items_bought = sum(products_list.values())

            
                purchase_data = [
                    '',
                    str(purchase_product.supplier),  
                    purchase_product.date_added.strftime('%Y-%m-%d %H:%M'),  
                    ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                    purchase_product.total,
                    total_items_bought
                ]

                ws.append(purchase_data)

            
            total_row = ['Total General:', total_suppliers, '', '', total_costos, total_items_comprados]
            ws.append(total_row)

            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.column in [2, 3, 4]:
                        cell.alignment = Alignment(horizontal='left')
                    elif cell.column in [5, 6, 7]:
                        cell.alignment = Alignment(horizontal='right')

            
            total_general_cell = ws['A' + str(ws.max_row)]
            total_general_cell.alignment = Alignment(horizontal='left')

            
            total_suppliers_cell = ws['B' + str(ws.max_row)]
            total_suppliers_cell.alignment = Alignment(horizontal='right')

            total_costos_cell = ws['E' + str(ws.max_row)]
            total_costos_cell.alignment = Alignment(horizontal='right')

            total_items_comprados_cell = ws['F' + str(ws.max_row)]
            total_items_comprados_cell.alignment = Alignment(horizontal='right')

            current_date = datetime.now()
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=reporte_compras_general_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

            wb.save(response)

            return response
        else:
            return HttpResponseBadRequest("Formulario no válido")



class ExcelPurchaseYearView(FormView):
    form_class = YearReportForm

    def form_valid(self, form):
        year = form.cleaned_data['year']

        
        purchase_products = PurchaseProduct.objects.filter(date_added__year=year)

        total_suppliers = purchase_products.values('supplier').distinct().count()
        total_items_comprados = purchase_products.aggregate(total=Sum('qty'))['total']
        total_costos = purchase_products.aggregate(total=Sum('total'))['total']

        wb = Workbook()
        ws = wb.active

        ws.append([f"Reporte de Ventas: del {year}"])
        headers = ['Proveedor', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
        ws.append([''] + headers)

        for purchase_product in purchase_products:
            products_list = {purchase_product.product.name: purchase_product.qty}
            total_items_bought = sum(products_list.values())

            purchase_data = [
                '',
                str(purchase_product.supplier),
                purchase_product.date_added.strftime('%Y-%m-%d %H:%M'),
                ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                purchase_product.total,
                total_items_bought
            ]
            ws.append(purchase_data)

        total_row = ['Total General:', total_suppliers, '', '', total_costos, total_items_comprados]
        ws.append(total_row)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.column in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='left')
                elif cell.column in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal='right')

        total_general_cell = ws['A' + str(ws.max_row)]
        total_general_cell.alignment = Alignment(horizontal='left')

        total_suppliers_cell = ws['B' + str(ws.max_row)]
        total_suppliers_cell.alignment = Alignment(horizontal='right')

        total_costos_cell = ws['E' + str(ws.max_row)]
        total_costos_cell.alignment = Alignment(horizontal='right')

        total_items_comprados_cell = ws['F' + str(ws.max_row)]
        total_items_comprados_cell.alignment = Alignment(horizontal='right')

        current_date = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=reporte_compras_anual_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)

        return response



class ExcelPurchaseMonthView(FormView):
    form_class = MonthYearReportForm

    def form_valid(self, form):
        year = form.cleaned_data['year']
        month = form.cleaned_data['month']
        try:
            month_name = MONTH_NAMES[month - 1]
        except IndexError:
            return HttpResponseBadRequest("El mes proporcionado no es válido.")

        
        purchase_products = PurchaseProduct.objects.filter(date_added__year=year, date_added__month=month)

        total_suppliers = purchase_products.values('supplier').distinct().count()
        total_items_comprados = purchase_products.aggregate(total=Sum('qty'))['total']
        total_costos = purchase_products.aggregate(total=Sum('total'))['total']

        wb = Workbook()
        ws = wb.active
        ws.append([f"Reporte de Ventas: de {month_name} del {year}"])
        headers = ['Proveedor', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
        ws.append([''] + headers)

        for purchase_product in purchase_products:
            products_list = {purchase_product.product.name: purchase_product.qty}
            total_items_bought = sum(products_list.values())

            purchase_data = [
                '',
                str(purchase_product.supplier),
                purchase_product.date_added.strftime('%Y-%m-%d %H:%M'),
                ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                purchase_product.total,
                total_items_bought
            ]
            ws.append(purchase_data)

        total_row = ['Total General:', total_suppliers, '', '', total_costos, total_items_comprados]
        ws.append(total_row)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.column in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='left')
                elif cell.column in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal='right')

        total_general_cell = ws['A' + str(ws.max_row)]
        total_general_cell.alignment = Alignment(horizontal='left')

        total_suppliers_cell = ws['B' + str(ws.max_row)]
        total_suppliers_cell.alignment = Alignment(horizontal='right')

        total_costos_cell = ws['E' + str(ws.max_row)]
        total_costos_cell.alignment = Alignment(horizontal='right')

        total_items_comprados_cell = ws['F' + str(ws.max_row)]
        total_items_comprados_cell.alignment = Alignment(horizontal='right')

        current_date = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=reporte_compras_mensual_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)

        return response



class ExcelPurchaseDayView(FormView):
    form_class = DayMonthYearReportForm

    def form_valid(self, form):
        year = form.cleaned_data['year']
        month = form.cleaned_data['month']
        day = form.cleaned_data['day']

        try:
            month_name = MONTH_NAMES[month - 1]
        except IndexError:
            messages.error(self.request, "El mes proporcionado no es válido.")
            return HttpResponseBadRequest("El mes proporcionado no es válido.")

        if not is_valid_day(year, month, day):
            messages.error(self.request, "La fecha ingresada no es válida.")
            return HttpResponseBadRequest("La fecha ingresada no es válida.")

        
        purchase_products = PurchaseProduct.objects.filter(date_added__year=year, date_added__month=month, date_added__day=day)

        total_suppliers = purchase_products.values('supplier').distinct().count()
        total_items_comprados = purchase_products.aggregate(total=Sum('qty'))['total']
        total_costos = purchase_products.aggregate(total=Sum('total'))['total']

        wb = Workbook()
        ws = wb.active
        ws.append([f"Reporte de Ventas - Día: {day} de {month_name} del {year}"])
        headers = ['Proveedor', 'Fecha', 'Productos', 'Total', 'Cantidad Total de Ítems']
        ws.append([''] + headers)

        for purchase_product in purchase_products:
            products_list = {purchase_product.product.name: purchase_product.qty}
            total_items_bought = sum(products_list.values())

            purchase_data = [
                '',
                str(purchase_product.supplier),
                purchase_product.date_added.strftime('%Y-%m-%d %H:%M'),
                ', '.join([f"{product}: {quantity}" for product, quantity in products_list.items()]),
                purchase_product.total,
                total_items_bought
            ]
            ws.append(purchase_data)

        total_row = ['Total General:', total_suppliers, '', '', total_costos, total_items_comprados]
        ws.append(total_row)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.column in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='left')
                elif cell.column in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal='right')

        total_general_cell = ws['A' + str(ws.max_row)]
        total_general_cell.alignment = Alignment(horizontal='left')

        total_suppliers_cell = ws['B' + str(ws.max_row)]
        total_suppliers_cell.alignment = Alignment(horizontal='right')

        total_costos_cell = ws['E' + str(ws.max_row)]
        total_costos_cell.alignment = Alignment(horizontal='right')

        total_items_comprados_cell = ws['F' + str(ws.max_row)]
        total_items_comprados_cell.alignment = Alignment(horizontal='right')

        current_date = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=reporte_compras_diario_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)

        return response
