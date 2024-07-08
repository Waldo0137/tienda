from purchase.models import *  # Importa el modelo PurchaseProduct
from pos.models import *
from inventory.models import *
from report.forms import *

import io
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
import uuid

from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render
from django.template.loader import get_template, render_to_string
from django.utils import timezone
from django.views import View
from django.views.generic.edit import FormView

import openpyxl
from openpyxl import Workbook


class SupplierExcelView(View):
    def get(self, request, *args, **kwargs):

        suppliers = Supplier.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Proveedores"


        headers = ['Proveedor', 'Información de contacto', 'Fecha de registro']
        ws.append(headers)


        for supplier in suppliers:
            supplier_data = [
                supplier.name,
                supplier.contact_info,
                supplier.date_added.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(supplier_data)


        current_date = datetime.now()
        filename = f"lista_proveedores_{current_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)

        return response

class SupplierProductExcelView(View):
    def get(self, request):
        suppliers = Supplier.objects.prefetch_related('purchaseproduct_set__product').all()

        
        wb = Workbook()
        ws = wb.active
        ws.title = "Proveedores y Productos"

        
        headers = ['Proveedor', 'Producto', 'Costo', 'Cantidad', 'Fecha de Adquisición']
        ws.append(headers)

        
        for supplier in suppliers:
            purchase_products = supplier.purchaseproduct_set.all()
            for purchase in purchase_products:
                product_data = [
                    supplier.name,
                    purchase.product.name if purchase.product else 'N/A',
                    purchase.cost,
                    purchase.qty,
                    purchase.date_added.strftime('%Y-%m-%d %H:%M:%S')
                ]
                ws.append(product_data)

        
        current_date = datetime.now()
        filename = f"lista_proveedores_productos_{current_date.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        
        wb.save(response)

        return response


class ProductExcelView(View):
    def get(self, request, *args, **kwargs):
        products = Products.objects.all().order_by('name')


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"


        headers = ['Nombre', 'Descripción', 'Fecha agregado']
        ws.append(headers)


        for product in products:
            product_data = [
                product.name,
                product.description,
                product.date_added.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            ws.append(product_data)


        current_date = datetime.now()
        filename = f"lista_productos_{current_date.strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'


        wb.save(response)

        return response

class ProductQtyExcelView(View):
    def get(self, request, *args, **kwargs):
        products = Products.objects.all().order_by('name')


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

    
        headers = ['Nombre', 'Descripción', 'Fecha de agregado', 'Cantidad', 'Precio', 'Costo']
        ws.append(headers)


        for product in products:
            product_data = [
                product.name,
                product.description,
                product.date_added.strftime('%Y-%m-%d %H:%M:%S'),
                product.cantidad,
                product.price,
                product.cost,
            ]
            ws.append(product_data)


        current_date = datetime.now()
        filename = f"lista_productos_detalles_{current_date.strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'


        wb.save(response)

        return response


class MixExcelSalesDayView(FormView):
    form_class = DayForm

    def form_valid(self, form):
        year = int(form.cleaned_data['year'])
        month = int(form.cleaned_data['month'])
        day = int(form.cleaned_data['day'])

    
        start_date = datetime(year, month, day, 3, 0, 0)
        end_date = start_date + timedelta(days=1)
        date_screen = datetime(year, month ,day)
        month_name = MONTH_NAMES[month - 1]
        
        day_name_english = start_date.strftime('%A')  
        day_name = DAYS_OF_WEEK[day_name_english]  

    
        if not self.is_valid_day(year, month, day):
            messages.error(self.request, "La fecha ingresada no es válida.")
            return self.form_invalid(form)

    
        sales = Sales.objects.filter(date_added__gte=start_date, date_added__lt=end_date)
        total_clientes = sales.values('id').distinct().count()
        total_items_vendidos = salesItems.objects.filter(sale__in=sales).aggregate(total=Sum('qty'))['total']
        total_ingresos = sales.aggregate(total=Sum('grand_total'))['total']

        sale_details = []
        total_net_profit = Decimal(0)
        for sale in sales:
            items = salesItems.objects.filter(sale=sale)
            products_list = {}
            net_profit_total = Decimal(0)
            for item in items:
                product_name = item.product.name
                products_list[product_name] = {
                    'qty': item.qty,
                    'price': item.product.price,
                    'cost': item.product.cost,
                }
                item_profit = Decimal(item.qty) * (item.product.price - item.product.cost)
                net_profit_total += item_profit
            total_net_profit += net_profit_total
            sale_details.append({
                'cliente': sale.cliente,
                'date_added': sale.date_added,
                'products_list': products_list,
                'grand_total': sale.grand_total,
                'net_profit': net_profit_total
            })

        return self.generate_excel(sale_details, total_clientes, total_items_vendidos, total_ingresos, total_net_profit, year, month_name, day, day_name, date_screen)

    def generate_excel(self, sale_details, total_clientes, total_items_vendidos, total_ingresos, total_net_profit, year, month_name, day, day_name, date_screen):
        # Crear el archivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Ventas Diario"

        # Escribir los encabezados del reporte
        sheet.append(["Fecha", "Cliente", "Producto", "Cantidad", "Precio", "Costo", "Ganancia Neta"])
        for sale in sale_details:
            for product, details in sale['products_list'].items():
                sheet.append([
                    sale['date_added'].strftime('%Y-%m-%d %H:%M:%S'),
                    sale['cliente'],
                    product,
                    details['qty'],
                    details['price'],
                    details['cost'],
                    sale['net_profit']
                ])

    
        sheet.append([])
        sheet.append(["Total Clientes", total_clientes])
        sheet.append(["Total Items Vendidos", total_items_vendidos])
        sheet.append(["Total Ingresos", total_ingresos])
        sheet.append(["Ganancia Neta Total", total_net_profit])
        
        sheet.append(["Fecha para la Solicitud", date_screen.strftime('%Y-%m-%d')])
    
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=reporte_cierreventas_diario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
    
        buffer = io.BytesIO()
        workbook.save(buffer)
        response.write(buffer.getvalue())
        
        return response

    def is_valid_day(self, year, month, day):
        try:
            datetime(year, month, day)
            return True
        except ValueError:
            return False



class MixTramoExcelSalesDayView(FormView):
    form_class = DayTramoForm

    def form_valid(self, form):
        start_year = int(form.cleaned_data['start_year'])
        start_month = int(form.cleaned_data['start_month'])
        start_day = int(form.cleaned_data['start_day'])
        end_year = int(form.cleaned_data['end_year'])
        end_month = int(form.cleaned_data['end_month'])
        end_day = int(form.cleaned_data['end_day'])
        
    
        if not self.is_valid_date_range(start_year, start_month, start_day, end_year, end_month, end_day):
            messages.error(self.request, "La fecha de inicio no puede ser mayor que la fecha de fin.")
            return self.form_invalid(form)

    
        start_date = datetime(start_year, start_month, start_day, 3, 0, 0)
        end_date = datetime(end_year, end_month, end_day, 3, 0, 0) + timedelta(days=1)
        start_screen = datetime(start_year, start_month, start_day)
        end_date_display = datetime(end_year, end_month, end_day)
    
        if not self.is_valid_day(start_year, start_month, start_day) or not self.is_valid_day(end_year, end_month, end_day):
            messages.error(self.request, "Una de las fechas ingresadas no es válida.")
            return self.form_invalid(form)
        
        day_name_start_english = start_date.strftime('%A')  
        day_name_start = DAYS_OF_WEEK[day_name_start_english]
        day_name_end_english = end_date_display.strftime('%A')
        day_name_end = DAYS_OF_WEEK[day_name_end_english]  
        
    
        sales = Sales.objects.filter(date_added__gte=start_date, date_added__lt=end_date)
        total_clientes = sales.values('id').distinct().count()
        total_items_vendidos = salesItems.objects.filter(sale__in=sales).aggregate(total=Sum('qty'))['total']
        total_ingresos = sales.aggregate(total=Sum('grand_total'))['total']

        sale_details = []
        total_net_profit = Decimal(0)
        for sale in sales:
            items = salesItems.objects.filter(sale=sale)
            products_list = {}
            net_profit_total = Decimal(0)
            for item in items:
                product_name = item.product.name
                products_list[product_name] = {
                    'qty': item.qty,
                    'price': item.product.price,
                    'cost': item.product.cost,
                }
                item_profit = Decimal(item.qty) * (item.product.price - item.product.cost)
                net_profit_total += item_profit
            total_net_profit += net_profit_total
            sale_details.append({
                'cliente': sale.cliente,
                'date_added': sale.date_added,
                'products_list': products_list,
                'grand_total': sale.grand_total,
                'net_profit': net_profit_total
            })

        return self.generate_excel(sale_details, total_clientes, total_items_vendidos, total_ingresos, total_net_profit, start_year, start_month, start_day, end_year, end_month, end_day, day_name_start, day_name_end, start_screen, end_date_display)

    def generate_excel(self, sale_details, total_clientes, total_items_vendidos, total_ingresos, total_net_profit, start_year, start_month, start_day, end_year, end_month, end_day, day_name_start, day_name_end, start_screen, end_date_display):
    
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Ventas Diario"

    
        sheet.append(["Fecha", "Cliente", "Producto", "Cantidad", "Precio", "Costo", "Ganancia Neta"])
        for sale in sale_details:
            for product, details in sale['products_list'].items():
                sheet.append([
                    sale['date_added'].strftime('%Y-%m-%d %H:%M:%S'),
                    sale['cliente'],
                    product,
                    details['qty'],
                    details['price'],
                    details['cost'],
                    sale['net_profit']
                ])

    
        sheet.append([])
        sheet.append(["Total Clientes", total_clientes])
        sheet.append(["Total Items Vendidos", total_items_vendidos])
        sheet.append(["Total Ingresos", total_ingresos])
        sheet.append(["Ganancia Neta Total", total_net_profit])
        sheet.append(["Fecha Inicio", start_screen.strftime('%Y-%m-%d')])
        sheet.append(["Fecha Fin", end_date_display.strftime('%Y-%m-%d')])

    
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=reporte_cierreventas_tramo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
    
        buffer = io.BytesIO()
        workbook.save(buffer)
        response.write(buffer.getvalue())
        
        return response

    def is_valid_day(self, year, month, day):
        try:
            datetime(year, month, day)
            return True
        except ValueError:
            return False

    def is_valid_date_range(self, start_year, start_month, start_day, end_year, end_month, end_day):
        start_date = datetime(start_year, start_month, start_day)
        end_date = datetime(end_year, end_month, end_day)
        return start_date <= end_date
